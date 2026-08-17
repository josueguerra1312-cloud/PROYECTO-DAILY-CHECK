from __future__ import annotations

import csv
from datetime import timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Flight, GroundWindow
from .normalization import combine, normalize_registration, normalize_text

REQUIRED_HEADERS = {
    "LEG DEPT DATE",
    "AC REG NUMBER",
    "FLT NUM",
    "DEP",
    "STD LT",
    "STA LT",
    "DST",
}


def _header_map(sheet) -> dict[str, int]:
    mapping = {
        normalize_text(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }
    missing = REQUIRED_HEADERS - mapping.keys()
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(sorted(missing))}")
    return mapping


def read_flights(path: str | Path, sheet_name: str = "VUELOS") -> list[Flight]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja {sheet_name!r}")
    sheet = workbook[sheet_name]
    columns = _header_map(sheet)
    flights: list[Flight] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        def value(header: str):
            return row[columns[header] - 1]

        if not value("AC REG NUMBER"):
            continue
        try:
            departure = combine(value("LEG DEPT DATE"), value("STD LT"))
            arrival = combine(value("LEG DEPT DATE"), value("STA LT"))
        except ValueError:
            continue
        if arrival < departure:
            arrival += timedelta(days=1)

        flights.append(
            Flight(
                operation_date=departure,
                registration=normalize_registration(value("AC REG NUMBER")),
                flight_number=normalize_text(value("FLT NUM")),
                origin=normalize_text(value("DEP")),
                scheduled_departure=departure,
                scheduled_arrival=arrival,
                destination=normalize_text(value("DST")),
                source_row=row_number,
            )
        )
    return flights


def _duration_text(value: timedelta | None) -> str:
    if value is None:
        return ""
    minutes = int(value.total_seconds() // 60)
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def write_xlsx(windows: Iterable[GroundWindow], path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SecuenciaGDL"
    headers = [
        "A/C", "FL ARR", "FECHA ARR", "ARR", "FL DEPT", "FECHA DEPT",
        "DEPT", "TIEMPO TIERRA", "TIPO", "OBSERVACION",
    ]
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in windows:
        sheet.append([
            item.registration,
            item.arrival_flight,
            item.arrival_at.date(),
            item.arrival_at.time(),
            item.departure_flight or "",
            item.departure_at.date() if item.departure_at else "",
            item.departure_at.time() if item.departure_at else "SPARE",
            _duration_text(item.ground_time),
            item.status,
            item.note,
        ])

    for cell in sheet["C"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for column in ("D", "G"):
        for cell in sheet[column][1:]:
            if cell.value != "SPARE":
                cell.number_format = "hh:mm"
    widths = [14, 11, 13, 10, 11, 13, 10, 16, 18, 35]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def write_csv(windows: Iterable[GroundWindow], path: str | Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow([
            "A/C", "FL ARR", "FECHA ARR", "ARR", "FL DEPT", "FECHA DEPT",
            "DEPT", "TIEMPO TIERRA", "TIPO", "OBSERVACION",
        ])
        for item in windows:
            writer.writerow([
                item.registration,
                item.arrival_flight,
                item.arrival_at.strftime("%d/%m/%Y"),
                item.arrival_at.strftime("%H:%M"),
                item.departure_flight or "",
                item.departure_at.strftime("%d/%m/%Y") if item.departure_at else "",
                item.departure_at.strftime("%H:%M") if item.departure_at else "SPARE",
                _duration_text(item.ground_time),
                item.status,
                item.note,
            ])
